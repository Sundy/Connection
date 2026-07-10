from pathlib import Path


def extract_text_from_document(file_path: str, file_type: str) -> str:
    path = Path(file_path)
    suffix = path.suffix.lower()
    if suffix in {".txt", ".md", ".csv"}:
        return path.read_text(encoding="utf-8", errors="ignore").strip()
    if suffix == ".pdf":
        return _extract_pdf(path)
    if suffix == ".docx":
        return _extract_docx(path)
    if suffix in {".xlsx", ".xlsm"}:
        return _extract_xlsx(path)
    return ""


def _extract_pdf(path: Path) -> str:
    try:
        from pypdf import PdfReader
    except ImportError:
        return ""

    reader = PdfReader(str(path))
    parts = [page.extract_text() or "" for page in reader.pages]
    return "\n".join(part for part in parts if part.strip()).strip()


def _extract_docx(path: Path) -> str:
    try:
        from docx import Document
    except ImportError:
        return ""

    document = Document(str(path))
    return "\n".join(paragraph.text for paragraph in document.paragraphs if paragraph.text.strip()).strip()


def _extract_xlsx(path: Path) -> str:
    try:
        from openpyxl import load_workbook
    except ImportError:
        return ""

    workbook = load_workbook(str(path), data_only=True, read_only=True)
    rows: list[str] = []
    for sheet in workbook.worksheets:
        rows.append(f"[{sheet.title}]")
        for row in sheet.iter_rows(values_only=True):
            values = [str(value) for value in row if value is not None and str(value).strip()]
            if values:
                rows.append(" ".join(values))
    return "\n".join(rows).strip()
